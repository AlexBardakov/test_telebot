import openpyxl

from datetime import datetime


def date_catalog_refresh():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['A2'].value = datetime.now().strftime('%d.%m.%Y')
    ws.save('Database.xlsx')


def melted_cheese_status_off():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['F8'].value = 'нет'
    ws.save('Database.xlsm')


def melted_cheese_status_on():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['F8'].value = 'да'
    ws.save('Database.xlsm')


def desserts_status_on():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['K8'].value = 'да'
    ws.save('Database.xlsm')


def desserts_status_off():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['K8'].value = 'нет'
    ws.save('Database.xlsm')


def pastafilata_status_off():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['A8'].value = 'нет'
    ws.save('Database.xlsm')


def pastafilata_status_on():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['A8'].value = 'да'
    ws.save('Database.xlsm')


def pastafilata_burrata_status_on():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['A8'].value = 'да(б)'
    ws.save('Database.xlsm')


def pastafilata_stracha_status_on():
    ws = openpyxl.load_workbook('Database.xlsx')
    wb = ws.active
    wb['A8'].value = 'да(с)'
    ws.save('Database.xlsm')
